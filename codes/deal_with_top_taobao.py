#coding:utf-8
import os
import csv
import pandas as pd
from openpyxl import Workbook 

# 每期仅需要修改月份信息
curmonth = '2017-08'

working_space = '/Users/Alas/Documents/TD_handover/top_taobao/'
# 存放原始数据的文件夹
path_source_file = working_space + 'data_source/' + curmonth + '/' + curmonth + '/'
date_dir_list = os.listdir(path_source_file)
path_to_write = working_space + 'results/' + curmonth + '/top_taobao_excel/'

# 将date_dir_list列表中的隐藏文件'.DS_Store'删除
# 为了通用删除(可能有多个重复值的)不必要的项, 没有直接删除列表第一项而是循环查找
#print dir_list
for item in ['.DS_Store',]:
    for i,d in enumerate(date_dir_list):
        if item in d:
	    del date_dir_list[i]
#print date_dir_list

# 控制是在哪一天的文件夹下
#for date in ['20170428',]: ##### 选择一天
for date in date_dir_list:
    print date

    path_dir_oneday = path_source_file + '/' + date
    category_list = os.listdir(path_dir_oneday)
    #for info in category_list:
	#print info
    #print '\n'

    # 删除日期文件夹下淘宝排行下文件列表中的系统文件
    # 或者称作提取文件名包含'revised_'子串的文件
    revised_category_list = []
    for d in category_list:
	if 'revised_' in d:
	    revised_category_list.append(d)
    #print date, revised_category_list

#    for c in revised_category_list[-2:-1]: ##### 选择一个大品类文件
#    for c in category_list:
    for c in revised_category_list:
	print c, '\n'

	df_category = pd.read_csv(path_dir_oneday + '/' + c, header=None)
        df_category.fillna('N/A', inplace=True)
        # 品类 子品类 排行榜类型
        # 1排名, 2关键词, 3参考价(占位,有意缺失), 4.1关注指数, 4.2关注比例显示条长度, 
        # 5.1升降位次, 5.2位次是否升降, 6.1升降幅度, 6.2幅度是否升降
        # 1升0降2持平
        list_columns = ['NavName', 'SubNavName', 'SubNavTypesName', 'Rank', 'Keyword', 'ReferencePrice', 
                'WatchedIndex', 'WI_RatioBar', 'RankChange', 'RC_YorN', 'RangeRankChange', 'RRC_YorN',]
	df_category.columns = list_columns
	#print df_category
	#print df_category['NavName'].unique()
	
	# 为8个大品类Excel文件新建6张榜单表
	wb = Workbook()
	ws = wb.active
	ws.title = u'品牌热门排行'
        #ws.append(list_columns)
	ws = wb.create_sheet(u'品牌上升榜')
        #ws.append(list_columns)
	ws = wb.create_sheet(u'搜索热门排行')
        #ws.append(list_columns)
	ws = wb.create_sheet(u'搜索上升榜')
        #ws.append(list_columns)
	ws = wb.create_sheet(u'销售热门排行')
        #ws.append(list_columns)
	ws = wb.create_sheet(u'销售上升榜')
        #ws.append(list_columns)
	
	# df_NavName 代表大品类(文件)下的某一品类
#	for nav_name in df_category['NavName'].unique()[:1]: ##### 选择一个品类
	for nav_name in df_category['NavName'].unique():
	    #print nav_name, '*'*10
	    #print '*'*20
	    df_NavName = df_category[df_category['NavName']==nav_name]
	    #print df_NavName
	    
	    # df_SubNavName 代表某一品类下的小品类
#            for sub_nav_name in df_NavName['SubNavName'].unique()[4:5]: ##### 选择一个小品类
	    for sub_nav_name in df_NavName['SubNavName'].unique():
		#print sub_nav_name
		df_SubNavName = df_NavName[df_NavName['SubNavName']==sub_nav_name]
		#print df_SubNavName

		# df_SubNavTypesName 代表某个小品类下的排行榜类型
#		for sub_nav_types_name in df_SubNavName['SubNavTypesName'].unique()[-1:]: ##### 选择一个榜单
		for sub_nav_types_name in df_SubNavName['SubNavTypesName'].unique():
		    #print sub_nav_types_name
		    df_SubNavTypesName = df_SubNavName[df_SubNavName['SubNavTypesName']==sub_nav_types_name]
		    #print df_SubNavTypesName
		
		    # 来进行排序
		    #print df_SubNavTypesName
		    sorted_df_SubNavTypesName = df_SubNavTypesName.sort_values(by='Rank')
		    #print sorted_df_SubNavTypesName
		    #print sorted_df_SubNavTypesName[:2]
		    #print list(sorted_df_SubNavTypesName[0:1].values[0])
		    #print list(sorted_df_SubNavTypesName[1:2].values[0])

                    #print sorted_df_SubNavTypesName.head()
                    #print nav_name, sub_nav_name, sub_nav_types_name

		    # 将排序后的结果存入Excel文件
		    for i in range(len(sorted_df_SubNavTypesName)):
			to_write_strings = list(sorted_df_SubNavTypesName[i:i+1].values[0])
                        #print to_write_strings
			# 不用每一个元素都转换为字符串
			#to_write_strings = [str(i) for i in to_write_strings]
			#if '品牌热门' in sub_nav_types_name:
			#    for s in to_write_strings:
			#	print s,
			#    print '\r'
			if sub_nav_types_name == '品牌热门排行':
			    ws = wb['品牌热门排行']
			    ws.append(to_write_strings)
			elif sub_nav_types_name == '品牌上升榜':
			    ws = wb['品牌上升榜']
			    ws.append(to_write_strings)
			elif sub_nav_types_name == '搜索热门排行':
			    ws = wb['搜索热门排行']
			    ws.append(to_write_strings)
			elif sub_nav_types_name == '搜索上升榜':
			    ws = wb['搜索上升榜']
			    ws.append(to_write_strings)
			elif sub_nav_types_name == '销售热门排行':
			    ws = wb['销售热门排行']
			    ws.append(to_write_strings)
			elif sub_nav_types_name == '销售上升榜':
			    ws = wb['销售上升榜']
			    ws.append(to_write_strings)
			else:
			    continue 
	    wb.save(path_to_write + '淘宝排行' + '_' + str(c[8:-4]) + '_' + str(date) + '.xlsx')


